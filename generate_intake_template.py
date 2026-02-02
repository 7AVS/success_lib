#!/usr/bin/env python3
"""
generate_intake_template.py - Generate Success Library Intake Template (Schema v2.0)

Creates an Excel template with all required sheets and columns for metric intake.

Usage:
    python3 generate_intake_template.py

Output:
    intake/template/intake_template_v2.xlsx
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install --user openpyxl")
    exit(1)

from pathlib import Path

OUTPUT_PATH = Path("intake/template/intake_template_v2.xlsx")

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
DESC_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def setup_sheet(ws, columns, descriptions, example_data, widths):
    """Set up a sheet with headers, descriptions, and example row."""

    # Row 1: Headers
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Row 2: Descriptions
    for col_idx, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.fill = DESC_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = THIN_BORDER

    # Row 3: Example (as comment/guide)
    for col_idx, example in enumerate(example_data, 1):
        cell = ws.cell(row=3, column=col_idx, value=example)
        cell.fill = EXAMPLE_FILL
        cell.alignment = Alignment(wrap_text=True)
        cell.border = THIN_BORDER

    # Set column widths
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header rows
    ws.freeze_panes = 'A4'


def create_business_sheet(wb):
    """Create the Business Metadata sheet."""
    ws = wb.create_sheet("Business Metadata")

    columns = [
        "action",
        "metric_id",
        "metric_name",
        "product",
        "line_of_business",
        "metric_type",
        "pillar",
        "business_definition",
        "source_tables",
        "staged_source",
        "grain",
        "owner",
        "campaigns_using",
    ]

    descriptions = [
        "new or update",
        "Unique ID: {PRODUCT}_{TYPE}_{SEQ}",
        "Human-readable name",
        "VVD, CC, MTG, PL, etc.",
        "Debit Cards, Lending, etc.",
        "Acquisition, Activation, Usage, Retention, Provisioning",
        "Conversion, Engagement, Retention, Profitability",
        "Plain English description of what success means",
        "Comma-separated source tables",
        "Staged/processed table if applicable",
        "Client, Account, or Transaction",
        "Team/person responsible",
        "Comma-separated campaign codes",
    ]

    example_data = [
        "new",
        "VVD_ACQ_001",
        "Card Acquisition",
        "VVD",
        "Debit Cards",
        "Acquisition",
        "Conversion",
        "Client acquired a new VVD card with active status",
        "DDWTA_VISA_DR_CRD",
        "",
        "Client",
        "Marketing Analytics",
        "VCN, VDA",
    ]

    widths = [10, 15, 20, 10, 15, 15, 15, 50, 30, 20, 12, 20, 20]

    setup_sheet(ws, columns, descriptions, example_data, widths)
    return ws


def create_technical_sheet(wb):
    """Create the Technical Metadata sheet."""
    ws = wb.create_sheet("Technical")

    columns = [
        "metric_id",
        "complexity",
        "source",
        "table_path",
        "date_field",
        "client_field",
        "partition_field",
        "client_extraction_logic",
        "add_card_type",
        "filter_1_field",
        "filter_1_condition",
        "filter_1_value",
        "filter_2_field",
        "filter_2_condition",
        "filter_2_value",
        "filter_3_field",
        "filter_3_condition",
        "filter_3_value",
        "filter_4_field",
        "filter_4_condition",
        "filter_4_value",
        "join_table",
        "join_type",
        "join_condition",
    ]

    descriptions = [
        "Must match Business sheet",
        "SIMPLE, TRANSFORM, MULTI_TABLE",
        "HIVE, EDW, TERADATA",
        "Full HDFS path (for HIVE)",
        "Column for success date",
        "Output column for client ID",
        "Partition column for optimization",
        "SQL expression if CLNT_NO not direct",
        "TRUE/FALSE",
        "Filter field name",
        "IN, =, >, <, NOT_NULL",
        "Filter value(s) comma-separated",
        "Filter field name",
        "Condition",
        "Value(s)",
        "Filter field name",
        "Condition",
        "Value(s)",
        "Filter field name",
        "Condition",
        "Value(s)",
        "Secondary table for MULTI_TABLE",
        "INNER, LEFT",
        "Join clause: A.key = B.key",
    ]

    example_data = [
        "VVD_ACQ_001",
        "SIMPLE",
        "HIVE",
        "/prod/sz/tsz/00050/data/DDWTA_VISA_DR_CRD/",
        "ISS_DT",
        "CLNT_NO",
        "CAPTR_DT",
        "",
        "TRUE",
        "STS_CD",
        "IN",
        "06,08",
        "SRVC_ID",
        "=",
        "36",
        "ISS_DT",
        "NOT_NULL",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    widths = [15, 12, 10, 50, 12, 12, 15, 30, 12, 15, 12, 15, 15, 12, 15, 15, 12, 15, 15, 12, 15, 30, 10, 30]

    setup_sheet(ws, columns, descriptions, example_data, widths)
    return ws


def create_measurement_sheet(wb):
    """Create the Measurement & Governance sheet."""
    ws = wb.create_sheet("Measurement")

    columns = [
        "metric_id",
        "default_window_days",
        "attribution_model",
        "status",
        "last_validated",
        "notes",
    ]

    descriptions = [
        "Must match Business sheet",
        "Measurement window in days (default: 90)",
        "FIRST_TOUCH, LAST_TOUCH, LINEAR",
        "PLACEHOLDER, ACTIVE, DEPRECATED, RETIRED",
        "YYYY-MM-DD of last validation",
        "Any additional notes",
    ]

    example_data = [
        "VVD_ACQ_001",
        "90",
        "FIRST_TOUCH",
        "ACTIVE",
        "2026-01-24",
        "Primary metric for acquisition campaigns",
    ]

    widths = [15, 20, 18, 15, 15, 50]

    setup_sheet(ws, columns, descriptions, example_data, widths)
    return ws


def create_instructions_sheet(wb):
    """Create the Instructions sheet."""
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    instructions = [
        ("Success Library Intake Template v2.0", True, 16),
        ("", False, 11),
        ("HOW TO USE THIS TEMPLATE:", True, 12),
        ("1. Fill in the Business Metadata sheet (required)", False, 11),
        ("2. Fill in the Technical sheet (required for engine integration)", False, 11),
        ("3. Fill in the Measurement sheet (optional, defaults provided)", False, 11),
        ("4. Save the file and place in intake/pending/", False, 11),
        ("5. Run: python3 excel_to_json.py", False, 11),
        ("", False, 11),
        ("SHEET DESCRIPTIONS:", True, 12),
        ("", False, 11),
        ("Business Metadata:", True, 11),
        ("- Captures what the metric measures and who owns it", False, 11),
        ("- Required fields: action, metric_id, metric_name, product, metric_type, pillar, owner", False, 11),
        ("", False, 11),
        ("Technical:", True, 11),
        ("- Captures how to execute the metric (table paths, filters, date fields)", False, 11),
        ("- Required for Vintage Engine integration", False, 11),
        ("- Complexity levels: SIMPLE (single table), TRANSFORM (client extraction), MULTI_TABLE (joins)", False, 11),
        ("", False, 11),
        ("Measurement:", True, 11),
        ("- Captures measurement window, attribution model, and governance status", False, 11),
        ("- Optional - defaults will be applied if not specified", False, 11),
        ("", False, 11),
        ("METRIC ID CONVENTION:", True, 12),
        ("{PRODUCT}_{TYPE}_{SEQUENCE}", False, 11),
        ("Examples: VVD_ACQ_001, CC_USG_002, MTG_RET_001", False, 11),
        ("", False, 11),
        ("FILTER CONDITIONS:", True, 12),
        ("IN - Multiple values: STS_CD IN ('06','08')", False, 11),
        ("= or EQ - Exact match: SRVC_ID = 36", False, 11),
        ("> or GT - Greater than: AMT1 > 0", False, 11),
        ("< or LT - Less than: DAYS_SINCE < 90", False, 11),
        ("NOT_NULL - Field is not null", False, 11),
        ("", False, 11),
        ("COMPLEXITY LEVELS:", True, 12),
        ("SIMPLE - Single table, CLNT_NO directly available", False, 11),
        ("TRANSFORM - Single table, CLNT_NO must be extracted (e.g., from card number)", False, 11),
        ("MULTI_TABLE - Requires joining multiple tables", False, 11),
        ("", False, 11),
        ("For detailed documentation, see: docs/08_INTAKE_GUIDE.md", False, 11),
    ]

    for row_idx, (text, is_bold, font_size) in enumerate(instructions, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=is_bold, size=font_size)

    ws.column_dimensions['A'].width = 100

    # Move Instructions to first position
    wb.move_sheet("Instructions", offset=-3)

    return ws


def main():
    print("Generating Success Library Intake Template v2.0...")

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_business_sheet(wb)
    create_technical_sheet(wb)
    create_measurement_sheet(wb)
    create_instructions_sheet(wb)

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    print(f"Created: {OUTPUT_PATH}")
    print()
    print("Template contains:")
    print("  - Instructions (overview and how-to)")
    print("  - Business Metadata (required)")
    print("  - Technical (for engine integration)")
    print("  - Measurement (optional governance)")
    print()
    print("Next steps:")
    print("  1. Open the template and fill in your metric data")
    print("  2. Save and place in intake/pending/")
    print("  3. Run: python3 excel_to_json.py")


if __name__ == "__main__":
    main()
