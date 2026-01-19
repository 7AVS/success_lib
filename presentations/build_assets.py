"""
Build Assets for Director Briefing Presentation

Generates Excel tables and text content for PowerPoint assembly.
Uses RBC brand colors.

Requirements:
    pip install openpyxl

Usage:
    python build_assets.py
"""

import os
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    exit(1)

# RBC Brand Colors
RBC = {
    'bright_blue': '0051A5',
    'dark_blue': '003168',
    'warm_yellow': 'FFC72C',
    'cool_white': 'E7EEF1',
    'ocean': '0091DA',
    'gray': '9EA2A2',
    'tundra': '07AFBF',
    'sunburst': 'FCA311',
}

# Paths
ASSETS_DIR = Path(__file__).parent / "assets"
TABLES_DIR = ASSETS_DIR / "tables"
TEXT_DIR = ASSETS_DIR / "text"

def create_styled_workbook():
    """Create workbook with RBC styling helpers."""
    wb = Workbook()
    return wb

def style_header(ws, row=1, cols=None):
    """Apply RBC header styling to a row."""
    header_fill = PatternFill(start_color=RBC['bright_blue'], end_color=RBC['bright_blue'], fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if cols is None:
        cols = range(1, ws.max_column + 1)

    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_rows(ws, start_row=2, end_row=None, cols=None):
    """Apply alternating row styling."""
    if end_row is None:
        end_row = ws.max_row
    if cols is None:
        cols = range(1, ws.max_column + 1)

    alt_fill = PatternFill(start_color=RBC['cool_white'], end_color=RBC['cool_white'], fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(start_row, end_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if (row - start_row) % 2 == 1:
                cell.fill = alt_fill

def auto_width(ws):
    """Auto-adjust column widths."""
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


# =============================================================================
# TABLE 1: Component Summary
# =============================================================================
def create_component_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Components"

    headers = ["Component", "Purpose", "Who Uses It", "Status"]
    data = [
        ["Metric Catalog", "Central index of all metrics", "System", "Complete"],
        ["Code Files", "SQL + PySpark calculation logic", "Analysts, Data Eng", "Structure complete"],
        ["HTML Interface", "Browse, search, view metrics", "Analysts", "Complete"],
        ["Intake Workflow", "Submit new metrics", "Analysts → Admin", "Complete"],
        ["Documentation", "Processes, decisions, guides", "Everyone", "Complete"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "01_components.xlsx")
    print("Created: 01_components.xlsx")


# =============================================================================
# TABLE 2: Before/After Comparison
# =============================================================================
def create_before_after_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Before_After"

    headers = ["Before (Today)", "After (Success Library)"]
    data = [
        ["Analysts write ad-hoc code for each campaign", "Reference governed, standardized definitions"],
        ["'Activation' means different things to different people", "One definition per metric, used everywhere"],
        ["No audit trail for calculations", "Full lineage: definition + code + version history"],
        ["Weeks spent recreating the same logic", "Look up, copy, run"],
        ["Results can't be compared across campaigns", "Consistent metrics enable cross-campaign analysis"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "02_before_after.xlsx")
    print("Created: 02_before_after.xlsx")


# =============================================================================
# TABLE 3: Data Architecture Trade-offs
# =============================================================================
def create_architecture_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Architecture"

    headers = ["Approach", "Governance", "Analytics", "AI-Ready"]
    data = [
        ["One table per product", "Medium", "Simple", "Good"],
        ["One unified table", "High", "Simple", "Good"],
        ["Generic + dimensions", "Medium", "Medium", "Good"],
        ["Granular metrics", "High", "Simple", "Best"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "03_architecture_tradeoffs.xlsx")
    print("Created: 03_architecture_tradeoffs.xlsx")


# =============================================================================
# TABLE 4: Decisions Needed
# =============================================================================
def create_decisions_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Decisions"

    headers = ["Decision Area", "Question", "Options", "Impact"]
    data = [
        ["Tech Stack", "What platform hosts the dataset?", "Snowflake / Hive / Redshift", "Schema design constraints"],
        ["Data Architecture", "One table or many?", "Per-product / Unified / Hybrid", "Query patterns, governance"],
        ["AI Integration", "How will AI access data?", "Direct query / Metadata + codegen", "Documentation requirements"],
        ["Governance", "Who maintains long-term?", "Dedicated role / Shared", "Resource allocation"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "04_decisions_needed.xlsx")
    print("Created: 04_decisions_needed.xlsx")


# =============================================================================
# TABLE 5: Timeline/Status
# =============================================================================
def create_status_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Status"

    headers = ["Phase", "Status", "Notes"]
    data = [
        ["Framework built", "Complete", "Infrastructure, templates, processes"],
        ["Content population", "In Progress", "Replacing placeholders with real logic"],
        ["Validation", "Not Started", "Testing against real data"],
        ["Adoption", "Not Started", "Training, rollout to analysts"],
        ["Automation", "Future", "Requires Data Engineering resources"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)

    # Color-code status column
    status_colors = {
        "Complete": "22C55E",      # Green
        "In Progress": "FFC72C",   # Yellow (RBC Warm Yellow)
        "Not Started": "9EA2A2",   # Gray
        "Future": "C1B5E0",        # Light purple
    }

    # Apply status colors
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=2).value
        if status in status_colors:
            ws.cell(row=row, column=2).fill = PatternFill(
                start_color=status_colors[status],
                end_color=status_colors[status],
                fill_type='solid'
            )

    auto_width(ws)

    wb.save(TABLES_DIR / "05_status_timeline.xlsx")
    print("Created: 05_status_timeline.xlsx")


# =============================================================================
# TABLE 6: What I Need From You
# =============================================================================
def create_asks_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Asks"

    headers = ["Ask", "Why It Matters"]
    data = [
        ["Tech stack decision", "Determines data architecture approach"],
        ["Connect me with Data Engineering", "Need to understand platform constraints"],
        ["AI/agent timeline", "How much to invest in metadata now?"],
        ["Resource alignment", "Who populates the metrics content?"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "06_asks.xlsx")
    print("Created: 06_asks.xlsx")


# =============================================================================
# TABLE 7: Initial Metrics
# =============================================================================
def create_metrics_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Metrics"

    headers = ["Product", "Metric ID", "Metric Name", "Type"]
    data = [
        ["VVD", "VVD_ACQ_001", "Virtual Visa Debit Acquisition", "Acquisition"],
        ["VVD", "VVD_ACT_001", "Virtual Visa Debit Activation", "Activation"],
        ["VVD", "VVD_USG_001", "Virtual Visa Debit Usage", "Usage"],
        ["VVD", "VVD_PRV_001", "Virtual Visa Debit Provisioning", "Provisioning"],
        ["CC", "CC_ACQ_001", "Credit Card Acquisition", "Acquisition"],
        ["MTG", "MTG_ACQ_001", "Mortgage Acquisition", "Acquisition"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "07_initial_metrics.xlsx")
    print("Created: 07_initial_metrics.xlsx")


# =============================================================================
# TABLE 8: Value Proposition
# =============================================================================
def create_value_table():
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Value"

    headers = ["Value Today", "Value Tomorrow (Automation)"]
    data = [
        ["Consistency - One definition per metric", "Automated calculation on schedule"],
        ["Speed - No time recreating logic", "Day 1 reporting for campaigns"],
        ["Auditability - Clear lineage", "Dashboard-ready standardized outputs"],
        ["Onboarding - New analysts find metrics quickly", "AI can query using metadata"],
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    style_header(ws)
    style_data_rows(ws)
    auto_width(ws)

    wb.save(TABLES_DIR / "08_value_proposition.xlsx")
    print("Created: 08_value_proposition.xlsx")


# =============================================================================
# MAIN
# =============================================================================
if __name__ == "__main__":
    # Ensure directories exist
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    TEXT_DIR.mkdir(parents=True, exist_ok=True)

    print("Building assets with RBC colors...")
    print(f"Output: {TABLES_DIR}")
    print()

    create_component_table()
    create_before_after_table()
    create_architecture_table()
    create_decisions_table()
    create_status_table()
    create_asks_table()
    create_metrics_table()
    create_value_table()

    print()
    print("Done! Excel tables created in assets/tables/")
