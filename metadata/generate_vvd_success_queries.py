#!/usr/bin/env python3
"""
Generate success_queries.xlsx - one tab per campaign mnemonic.
Each tab follows the source_to_target_mapping template:
  1. Source-to-target mapping table
  2. ORGANIC: Sample (10 rows) + Summary (by year-month)
  3. CAMPAIGN: Sample (10 rows) + Summary (by year-month)

All queries use EDW (Teradata) SQL syntax with PROC SQL / CONNECTION TO TERADATA pattern.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT      = Font(name="Calibri", bold=True, size=12)
SQL_FONT         = Font(name="Consolas", size=10)
LABEL_FONT       = Font(name="Calibri", bold=True, size=11)
MAP_HDR_FONT     = Font(name="Calibri", bold=True, size=11)
MAP_DATA_FONT    = Font(name="Calibri", size=11)

MAP_HDR_FILL     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FILL       = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL        = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
YELLOW_FILL      = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

WRAP_ALIGN       = Alignment(wrap_text=False, vertical="top")
THIN_BORDER      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COL_A_WIDTH      = 30
COL_B_WIDTH      = 45
COL_C_WIDTH      = 50

# ---------------------------------------------------------------------------
# Snap-date subquery
# ---------------------------------------------------------------------------
SNAP_SUBQUERY = (
    "(SELECT MAX(SNAP_DT) FROM DDWV01.VISA_DR_CRD_DLY "
    "WHERE SNAP_DT >= CURRENT_DATE - 7)"
)

# ---------------------------------------------------------------------------
# Source-to-target mapping rows per metric type
# ---------------------------------------------------------------------------

def _card_acq_mapping(mne):
    return [
        ("clnt_no",            "DDWV01.VISA_DR_CRD_DLY",  "CLNT_NO"),
        ("acct_no",            "DDWV01.VISA_DR_CRD_DLY",  ""),
        ("amount",             "",                          ""),
        ("event_mnemonic_cd",  "",                          mne),
        ("event_type_cd",      "",                          ""),
        ("event_attributes",   "DDWV01.VISA_DR_CRD_DLY",  "JSON(STS_CD, SRVC_ID)"),
        ("event_dt",           "DDWV01.VISA_DR_CRD_DLY",  "ISS_DT"),
        ("snap_dt",            "",                          "<GENERATED>"),
        ("job_id",             "",                          "<GENERATED>"),
    ]

def _card_actv_mapping(mne):
    return [
        ("clnt_no",            "DDWV01.VISA_DR_CRD_DLY",  "CLNT_NO"),
        ("acct_no",            "DDWV01.VISA_DR_CRD_DLY",  ""),
        ("amount",             "",                          ""),
        ("event_mnemonic_cd",  "",                          mne),
        ("event_type_cd",      "",                          ""),
        ("event_attributes",   "DDWV01.VISA_DR_CRD_DLY",  "JSON(STS_CD, SRVC_ID)"),
        ("event_dt",           "DDWV01.VISA_DR_CRD_DLY",  "ACTV_DT"),
        ("snap_dt",            "",                          "<GENERATED>"),
        ("job_id",             "",                          "<GENERATED>"),
    ]

def _card_usage_mapping(mne):
    return [
        ("clnt_no",            "DDWV05.CLNT_CRD_POS_LOG",  "CAST(SUBSTR(CLNT_CRD_NO, 7, 9) AS INTEGER)"),
        ("acct_no",            "DDWV05.CLNT_CRD_POS_LOG",  "CLNT_CRD_NO"),
        ("amount",             "DDWV05.CLNT_CRD_POS_LOG",  "AMT1"),
        ("event_mnemonic_cd",  "",                           mne),
        ("event_type_cd",      "",                           ""),
        ("event_attributes",   "DDWV05.CLNT_CRD_POS_LOG",  "JSON(TXN_TP, MSG_TP, SRVC_CD)"),
        ("event_dt",           "DDWV05.CLNT_CRD_POS_LOG",  "TXN_DT"),
        ("snap_dt",            "",                           "<GENERATED>"),
        ("job_id",             "",                           "<GENERATED>"),
    ]

def _wallet_mapping(mne):
    return [
        ("clnt_no",            "DDWV05.CLNT_CRD_POS_LOG",  "CAST(SUBSTR(CLNT_CRD_NO, 7, 9) AS INTEGER)"),
        ("acct_no",            "DDWV05.CLNT_CRD_POS_LOG",  "CLNT_CRD_NO"),
        ("amount",             "",                           ""),
        ("event_mnemonic_cd",  "",                           mne),
        ("event_type_cd",      "",                           ""),
        ("event_attributes",   "DDWV05.CLNT_CRD_POS_LOG + DL_DECMAN.TOKEN_LIST",
                                                             "JSON(TOKN_REQSTR_ID, TOKEN_WALLET_IND)"),
        ("event_dt",           "DDWV05.CLNT_CRD_POS_LOG",  "TXN_DT"),
        ("snap_dt",            "",                           "<GENERATED>"),
        ("job_id",             "",                           "<GENERATED>"),
    ]

def _imt_mapping(mne):
    return [
        ("clnt_no",            "DDWV01.EXT_CDS_CHNL_EVNT",  "CLNT_NO"),
        ("acct_no",            "DDWV01.EXT_CDS_CHNL_EVNT",  "AR_ID"),
        ("amount",             "DDWV01.EXT_CDS_CHNL_EVNT",  "EVNT_AMT_CAD"),
        ("event_mnemonic_cd",  "",                            mne),
        ("event_type_cd",      "",                            ""),
        ("event_attributes",   "DDWV01.EXT_CDS_CHNL_EVNT",  "JSON(CHNL_TYP_CD, EVNT_CRNCY_CD, ACTVY_TYP_CD, SRC_DTA_STORE_CD)"),
        ("event_dt",           "DDWV01.EXT_CDS_CHNL_EVNT",  "CAPTR_DT"),
        ("snap_dt",            "",                            "<GENERATED>"),
        ("job_id",             "",                            "<GENERATED>"),
    ]


# ===========================================================================
# Query builder functions
# Each returns (sample_sql, summary_sql) for organic and campaign
# ===========================================================================

# --- CARD ACQUISITION (VCN, VDA) -------------------------------------------

def card_acq_organic():
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        ISS_DT        AS SUCCESS_DT,
        STS_CD,
        SRVC_ID
    FROM DDWV01.VISA_DR_CRD_DLY
    WHERE STS_CD IN ('06','08')
      AND SRVC_ID = 36
      AND ISS_DT IS NOT NULL
      AND SNAP_DT = {SNAP_SUBQUERY}
      AND ISS_DT >= DATE '2025-01-01'
    ORDER BY ISS_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM ISS_DT)  AS yr,
        EXTRACT(MONTH FROM ISS_DT) AS mo,
        COUNT(DISTINCT CLNT_NO)    AS unique_clients,
        COUNT(*)                   AS total_events
    FROM DDWV01.VISA_DR_CRD_DLY
    WHERE STS_CD IN ('06','08')
      AND SRVC_ID = 36
      AND ISS_DT IS NOT NULL
      AND SNAP_DT = {SNAP_SUBQUERY}
      AND ISS_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def card_acq_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        A.CLNT_NO,
        A.ISS_DT            AS SUCCESS_DT,
        A.STS_CD,
        B.TACTIC_ID,
        B.TREATMT_STRT_DT,
        B.TREATMT_END_DT
    FROM DDWV01.VISA_DR_CRD_DLY        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON A.CLNT_NO = B.CLNT_NO
    WHERE A.STS_CD IN ('06','08')
      AND A.SRVC_ID = 36
      AND A.ISS_DT IS NOT NULL
      AND A.SNAP_DT = {SNAP_SUBQUERY}
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.ISS_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.ISS_DT >= DATE '2025-01-01'
    ORDER BY A.ISS_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM A.ISS_DT)  AS yr,
        EXTRACT(MONTH FROM A.ISS_DT) AS mo,
        COUNT(DISTINCT A.CLNT_NO)    AS unique_clients,
        COUNT(*)                     AS total_events
    FROM DDWV01.VISA_DR_CRD_DLY        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON A.CLNT_NO = B.CLNT_NO
    WHERE A.STS_CD IN ('06','08')
      AND A.SRVC_ID = 36
      AND A.ISS_DT IS NOT NULL
      AND A.SNAP_DT = {SNAP_SUBQUERY}
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.ISS_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.ISS_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- CARD ACTIVATION (VDT) -------------------------------------------------

def card_actv_organic():
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        ACTV_DT       AS SUCCESS_DT,
        STS_CD,
        SRVC_ID
    FROM DDWV01.VISA_DR_CRD_DLY
    WHERE STS_CD IN ('06','08')
      AND SRVC_ID = 36
      AND ISS_DT IS NOT NULL
      AND ACTV_DT IS NOT NULL
      AND SNAP_DT = {SNAP_SUBQUERY}
      AND ACTV_DT >= DATE '2025-01-01'
    ORDER BY ACTV_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM ACTV_DT)  AS yr,
        EXTRACT(MONTH FROM ACTV_DT) AS mo,
        COUNT(DISTINCT CLNT_NO)     AS unique_clients,
        COUNT(*)                    AS total_events
    FROM DDWV01.VISA_DR_CRD_DLY
    WHERE STS_CD IN ('06','08')
      AND SRVC_ID = 36
      AND ISS_DT IS NOT NULL
      AND ACTV_DT IS NOT NULL
      AND SNAP_DT = {SNAP_SUBQUERY}
      AND ACTV_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def card_actv_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        A.CLNT_NO,
        A.ACTV_DT           AS SUCCESS_DT,
        A.STS_CD,
        B.TACTIC_ID,
        B.TREATMT_STRT_DT,
        B.TREATMT_END_DT
    FROM DDWV01.VISA_DR_CRD_DLY        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON A.CLNT_NO = B.CLNT_NO
    WHERE A.STS_CD IN ('06','08')
      AND A.SRVC_ID = 36
      AND A.ISS_DT IS NOT NULL
      AND A.ACTV_DT IS NOT NULL
      AND A.SNAP_DT = {SNAP_SUBQUERY}
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.ACTV_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.ACTV_DT >= DATE '2025-01-01'
    ORDER BY A.ACTV_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM A.ACTV_DT)  AS yr,
        EXTRACT(MONTH FROM A.ACTV_DT) AS mo,
        COUNT(DISTINCT A.CLNT_NO)     AS unique_clients,
        COUNT(*)                      AS total_events
    FROM DDWV01.VISA_DR_CRD_DLY        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON A.CLNT_NO = B.CLNT_NO
    WHERE A.STS_CD IN ('06','08')
      AND A.SRVC_ID = 36
      AND A.ISS_DT IS NOT NULL
      AND A.ACTV_DT IS NOT NULL
      AND A.SNAP_DT = {SNAP_SUBQUERY}
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.ACTV_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.ACTV_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- CARD USAGE (VUI) ------------------------------------------------------

def card_usage_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CAST(SUBSTR(CLNT_CRD_NO, 7, 9) AS INTEGER)  AS CLNT_NO,
        TXN_DT            AS SUCCESS_DT,
        AMT1,
        TXN_TP,
        MSG_TP,
        SRVC_CD
    FROM DDWV05.CLNT_CRD_POS_LOG
    WHERE AMT1 > 0
      AND TXN_TP IN (10, 13)
      AND SUBSTR(CLNT_CRD_NO, 1, 5) = '45190'
      AND SRVC_CD = 36
      AND TXN_DT >= DATE '2025-01-01'
    ORDER BY TXN_DT DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM TXN_DT)  AS yr,
        EXTRACT(MONTH FROM TXN_DT) AS mo,
        COUNT(DISTINCT CAST(SUBSTR(CLNT_CRD_NO, 7, 9) AS INTEGER))
                                    AS unique_clients,
        COUNT(*)                    AS total_events,
        SUM(AMT1)                   AS total_amount
    FROM DDWV05.CLNT_CRD_POS_LOG
    WHERE AMT1 > 0
      AND TXN_TP IN (10, 13)
      AND SUBSTR(CLNT_CRD_NO, 1, 5) = '45190'
      AND SRVC_CD = 36
      AND TXN_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def card_usage_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CAST(SUBSTR(A.CLNT_CRD_NO, 7, 9) AS INTEGER)  AS CLNT_NO,
        A.TXN_DT             AS SUCCESS_DT,
        A.AMT1,
        A.TXN_TP,
        B.TACTIC_ID,
        B.TREATMT_STRT_DT,
        B.TREATMT_END_DT
    FROM DDWV05.CLNT_CRD_POS_LOG        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON CAST(SUBSTR(A.CLNT_CRD_NO, 7, 9) AS INTEGER) = B.CLNT_NO
    WHERE A.AMT1 > 0
      AND A.TXN_TP IN (10, 13)
      AND SUBSTR(A.CLNT_CRD_NO, 1, 5) = '45190'
      AND A.SRVC_CD = 36
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.TXN_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.TXN_DT >= DATE '2025-01-01'
    ORDER BY A.TXN_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM A.TXN_DT)  AS yr,
        EXTRACT(MONTH FROM A.TXN_DT) AS mo,
        COUNT(DISTINCT CAST(SUBSTR(A.CLNT_CRD_NO, 7, 9) AS INTEGER))
                                      AS unique_clients,
        COUNT(*)                      AS total_events,
        SUM(A.AMT1)                   AS total_amount
    FROM DDWV05.CLNT_CRD_POS_LOG        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON CAST(SUBSTR(A.CLNT_CRD_NO, 7, 9) AS INTEGER) = B.CLNT_NO
    WHERE A.AMT1 > 0
      AND A.TXN_TP IN (10, 13)
      AND SUBSTR(A.CLNT_CRD_NO, 1, 5) = '45190'
      AND A.SRVC_CD = 36
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.TXN_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.TXN_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- WALLET PROVISIONING (VUT, VAW) ----------------------------------------

def wallet_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CAST(SUBSTR(B.CLNT_CRD_NO, 7, 9) AS INTEGER)  AS CLNT_NO,
        B.TXN_DT                                        AS SUCCESS_DT,
        B.TOKN_REQSTR_ID,
        C.TOKEN_WALLET_IND
    FROM DDWV05.CLNT_CRD_POS_LOG  AS B
    INNER JOIN DL_DECMAN.TOKEN_LIST AS C
        ON B.TOKN_REQSTR_ID = C.TOKEN_ID
    WHERE B.AMT1 = 0
      AND SUBSTR(B.CLNT_CRD_NO, 1, 5)  = '45190'
      AND SUBSTR(B.VISA_DR_CRD_NO, 1, 5) = '45199'
      AND SUBSTR(B.TOKN_REQSTR_ID, 1, 1) > '0'
      AND B.POS_ENTR_MODE_CD_NON_EMV = '000'
      AND B.SRVC_CD = 36
      AND C.TOKEN_WALLET_IND = 'Y'
      AND B.TXN_DT >= DATE '2025-01-01'
    ORDER BY B.TXN_DT DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM B.TXN_DT)  AS yr,
        EXTRACT(MONTH FROM B.TXN_DT) AS mo,
        COUNT(DISTINCT CAST(SUBSTR(B.CLNT_CRD_NO, 7, 9) AS INTEGER))
                                      AS unique_clients,
        COUNT(*)                      AS total_events
    FROM DDWV05.CLNT_CRD_POS_LOG  AS B
    INNER JOIN DL_DECMAN.TOKEN_LIST AS C
        ON B.TOKN_REQSTR_ID = C.TOKEN_ID
    WHERE B.AMT1 = 0
      AND SUBSTR(B.CLNT_CRD_NO, 1, 5)  = '45190'
      AND SUBSTR(B.VISA_DR_CRD_NO, 1, 5) = '45199'
      AND SUBSTR(B.TOKN_REQSTR_ID, 1, 1) > '0'
      AND B.POS_ENTR_MODE_CD_NON_EMV = '000'
      AND B.SRVC_CD = 36
      AND C.TOKEN_WALLET_IND = 'Y'
      AND B.TXN_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def wallet_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CAST(SUBSTR(B.CLNT_CRD_NO, 7, 9) AS INTEGER)  AS CLNT_NO,
        B.TXN_DT                                        AS SUCCESS_DT,
        D.TACTIC_ID,
        D.TREATMT_STRT_DT,
        D.TREATMT_END_DT
    FROM DDWV05.CLNT_CRD_POS_LOG  AS B
    INNER JOIN DL_DECMAN.TOKEN_LIST AS C
        ON B.TOKN_REQSTR_ID = C.TOKEN_ID
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS D
        ON CAST(SUBSTR(B.CLNT_CRD_NO, 7, 9) AS INTEGER) = D.CLNT_NO
    WHERE B.AMT1 = 0
      AND SUBSTR(B.CLNT_CRD_NO, 1, 5)  = '45190'
      AND SUBSTR(B.VISA_DR_CRD_NO, 1, 5) = '45199'
      AND SUBSTR(B.TOKN_REQSTR_ID, 1, 1) > '0'
      AND B.POS_ENTR_MODE_CD_NON_EMV = '000'
      AND B.SRVC_CD = 36
      AND C.TOKEN_WALLET_IND = 'Y'
      AND SUBSTR(D.TACTIC_ID, 8, 3) = '{mne}'
      AND B.TXN_DT BETWEEN D.TREATMT_STRT_DT AND D.TREATMT_END_DT
      AND B.TXN_DT >= DATE '2025-01-01'
    ORDER BY B.TXN_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM B.TXN_DT)  AS yr,
        EXTRACT(MONTH FROM B.TXN_DT) AS mo,
        COUNT(DISTINCT CAST(SUBSTR(B.CLNT_CRD_NO, 7, 9) AS INTEGER))
                                      AS unique_clients,
        COUNT(*)                      AS total_events
    FROM DDWV05.CLNT_CRD_POS_LOG  AS B
    INNER JOIN DL_DECMAN.TOKEN_LIST AS C
        ON B.TOKN_REQSTR_ID = C.TOKEN_ID
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS D
        ON CAST(SUBSTR(B.CLNT_CRD_NO, 7, 9) AS INTEGER) = D.CLNT_NO
    WHERE B.AMT1 = 0
      AND SUBSTR(B.CLNT_CRD_NO, 1, 5)  = '45190'
      AND SUBSTR(B.VISA_DR_CRD_NO, 1, 5) = '45199'
      AND SUBSTR(B.TOKN_REQSTR_ID, 1, 1) > '0'
      AND B.POS_ENTR_MODE_CD_NON_EMV = '000'
      AND B.SRVC_CD = 36
      AND C.TOKEN_WALLET_IND = 'Y'
      AND SUBSTR(D.TACTIC_ID, 8, 3) = '{mne}'
      AND B.TXN_DT BETWEEN D.TREATMT_STRT_DT AND D.TREATMT_END_DT
      AND B.TXN_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- IMT (IRI) --------------------------------------------------------------

def imt_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        CAPTR_DT          AS SUCCESS_DT,
        EVNT_ID,
        EVNT_AMT_CAD,
        CHNL_TYP_CD,
        EVNT_CRNCY_CD
    FROM DDWV01.EXT_CDS_CHNL_EVNT
    WHERE ACTVY_TYP_CD = '031'
      AND CHNL_TYP_CD IN ('021','034')
      AND SRC_DTA_STORE_CD IN ('139','140')
      AND CAPTR_DT >= DATE '2025-01-01'
    ORDER BY CAPTR_DT DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM CAPTR_DT)  AS yr,
        EXTRACT(MONTH FROM CAPTR_DT) AS mo,
        COUNT(DISTINCT CLNT_NO)      AS unique_clients,
        COUNT(DISTINCT EVNT_ID)      AS unique_transactions,
        SUM(EVNT_AMT_CAD)            AS total_amt_cad
    FROM DDWV01.EXT_CDS_CHNL_EVNT
    WHERE ACTVY_TYP_CD = '031'
      AND CHNL_TYP_CD IN ('021','034')
      AND SRC_DTA_STORE_CD IN ('139','140')
      AND CAPTR_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def imt_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        A.CLNT_NO,
        A.CAPTR_DT           AS SUCCESS_DT,
        A.EVNT_ID,
        A.EVNT_AMT_CAD,
        A.CHNL_TYP_CD,
        B.TACTIC_ID,
        B.TREATMT_STRT_DT,
        B.TREATMT_END_DT
    FROM DDWV01.EXT_CDS_CHNL_EVNT        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON A.CLNT_NO = B.CLNT_NO
    WHERE A.ACTVY_TYP_CD = '031'
      AND A.CHNL_TYP_CD IN ('021','034')
      AND A.SRC_DTA_STORE_CD IN ('139','140')
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.CAPTR_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.CAPTR_DT >= DATE '2025-01-01'
    ORDER BY A.CAPTR_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM A.CAPTR_DT)  AS yr,
        EXTRACT(MONTH FROM A.CAPTR_DT) AS mo,
        COUNT(DISTINCT A.CLNT_NO)      AS unique_clients,
        COUNT(DISTINCT A.EVNT_ID)      AS unique_transactions,
        SUM(A.EVNT_AMT_CAD)            AS total_amt_cad
    FROM DDWV01.EXT_CDS_CHNL_EVNT        AS A
    INNER JOIN DG6V01.TACTIC_EVNT_IP_AR_HIST AS B
        ON A.CLNT_NO = B.CLNT_NO
    WHERE A.ACTVY_TYP_CD = '031'
      AND A.CHNL_TYP_CD IN ('021','034')
      AND A.SRC_DTA_STORE_CD IN ('139','140')
      AND SUBSTR(B.TACTIC_ID, 8, 3) = '{mne}'
      AND A.CAPTR_DT BETWEEN B.TREATMT_STRT_DT AND B.TREATMT_END_DT
      AND A.CAPTR_DT >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- O2P  (Overdraft — credit application tables) --------------------------

def _o2p_mapping(mne):
    return [
        ("clnt_no",            "DDWV01.CR_APP_CLNT_RELTN",  "CLNT_NO"),
        ("acct_no",            "DDWV01.CR_APP_PROD",         "PROD_APP_ACCT_TR_NO || APP_ACCT_NO  (derived AR_ID)"),
        ("amount",             "",                            ""),
        ("event_mnemonic_cd",  "",                            mne),
        ("event_type_cd",      "",                            ""),
        ("event_attributes",   "DDWV01.CR_APP_PROD",         "JSON(APPL_FOR_PROD_TYP, PROD_APP_STS_CD)"),
        ("event_dt",           "DDWV01.CR_APP_PROD",         "PROD_APP_COMPL_DT"),
        ("snap_dt",            "",                            "<GENERATED>"),
        ("job_id",             "",                            "<GENERATED>"),
    ]


def o2p_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        A.CLNT_NO,
        D.PROD_APP_COMPL_DT   AS SUCCESS_DT,
        D.APPL_FOR_PROD_TYP,
        D.PROD_APP_STS_CD,
        D.CR_APP_ID
    FROM DDWV01.CR_APP_CLNT_RELTN        AS A
    JOIN DDWV01.OVRL_CR_APP              AS B
        ON B.CR_APP_ID  = A.CR_APP_ID
        AND B.SYS_SRC_ID = A.SYS_SRC_ID
    JOIN DDWV01.CR_APP_CLNT_PROD_RELTN   AS C
        ON A.CR_APP_ID          = C.CR_APP_ID
        AND A.CR_APP_CLNT_SEQ_NO = C.CR_APP_CLNT_SEQ_NO
        AND A.SYS_SRC_ID        = C.SYS_SRC_ID
    JOIN DDWV01.CR_APP_PROD              AS D
        ON C.CR_APP_ID          = D.CR_APP_ID
        AND C.CR_APP_PROD_SEQ_NO = D.CR_APP_PROD_SEQ_NO
        AND C.SYS_SRC_ID        = D.SYS_SRC_ID
    JOIN DG6V01.ARNGMNT_OWN_HIST        AS E
        ON A.CLNT_NO = E.CLNT_NO
        AND E.MIF_SRVC_ID IN (4)
        AND E.MTH_END_DT = (SELECT MAX(MTH_END_DT) FROM DG6V01.ARNGMNT_OWN_HIST)
    WHERE D.APPL_FOR_PROD_TYP IN ('40','41','43')
      AND B.APP_TYP = 'P'
      AND D.PROD_APP_STS_CD IN (32,37,45,47,51,56,62)
      AND D.PROD_APP_COMPL_DT IS NOT NULL
      AND D.PROD_APP_COMPL_DT >= DATE '2025-01-01'
    ORDER BY D.PROD_APP_COMPL_DT DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM D.PROD_APP_COMPL_DT)  AS yr,
        EXTRACT(MONTH FROM D.PROD_APP_COMPL_DT) AS mo,
        COUNT(DISTINCT A.CLNT_NO)                AS unique_clients,
        COUNT(*)                                 AS total_events,
        D.APPL_FOR_PROD_TYP
    FROM DDWV01.CR_APP_CLNT_RELTN        AS A
    JOIN DDWV01.OVRL_CR_APP              AS B
        ON B.CR_APP_ID  = A.CR_APP_ID
        AND B.SYS_SRC_ID = A.SYS_SRC_ID
    JOIN DDWV01.CR_APP_CLNT_PROD_RELTN   AS C
        ON A.CR_APP_ID          = C.CR_APP_ID
        AND A.CR_APP_CLNT_SEQ_NO = C.CR_APP_CLNT_SEQ_NO
        AND A.SYS_SRC_ID        = C.SYS_SRC_ID
    JOIN DDWV01.CR_APP_PROD              AS D
        ON C.CR_APP_ID          = D.CR_APP_ID
        AND C.CR_APP_PROD_SEQ_NO = D.CR_APP_PROD_SEQ_NO
        AND C.SYS_SRC_ID        = D.SYS_SRC_ID
    JOIN DG6V01.ARNGMNT_OWN_HIST        AS E
        ON A.CLNT_NO = E.CLNT_NO
        AND E.MIF_SRVC_ID IN (4)
        AND E.MTH_END_DT = (SELECT MAX(MTH_END_DT) FROM DG6V01.ARNGMNT_OWN_HIST)
    WHERE D.APPL_FOR_PROD_TYP IN ('40','41','43')
      AND B.APP_TYP = 'P'
      AND D.PROD_APP_STS_CD IN (32,37,45,47,51,56,62)
      AND D.PROD_APP_COMPL_DT IS NOT NULL
      AND D.PROD_APP_COMPL_DT >= DATE '2025-01-01'
    GROUP BY 1, 2, D.APPL_FOR_PROD_TYP
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def o2p_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        A.CLNT_NO,
        D.PROD_APP_COMPL_DT   AS SUCCESS_DT,
        D.APPL_FOR_PROD_TYP,
        D.PROD_APP_STS_CD,
        T.TACTIC_ID,
        T.TREATMT_STRT_DT,
        T.TREATMT_END_DT
    FROM DG6V01.TACTIC_EVNT_IP_AR_HIST   AS T
    JOIN DDWV01.CR_APP_CLNT_RELTN        AS A
        ON T.CLNT_NO = A.CLNT_NO
    JOIN DDWV01.OVRL_CR_APP              AS B
        ON B.CR_APP_ID  = A.CR_APP_ID
        AND B.SYS_SRC_ID = A.SYS_SRC_ID
    JOIN DDWV01.CR_APP_CLNT_PROD_RELTN   AS C
        ON A.CR_APP_ID          = C.CR_APP_ID
        AND A.CR_APP_CLNT_SEQ_NO = C.CR_APP_CLNT_SEQ_NO
        AND A.SYS_SRC_ID        = C.SYS_SRC_ID
    JOIN DDWV01.CR_APP_PROD              AS D
        ON C.CR_APP_ID          = D.CR_APP_ID
        AND C.CR_APP_PROD_SEQ_NO = D.CR_APP_PROD_SEQ_NO
        AND C.SYS_SRC_ID        = D.SYS_SRC_ID
    JOIN DG6V01.ARNGMNT_OWN_HIST        AS E
        ON A.CLNT_NO = E.CLNT_NO
        AND E.MIF_SRVC_ID IN (4)
        AND E.MTH_END_DT = ADD_MONTHS(T.TREATMT_END_DT
            - (EXTRACT(DAY FROM T.TREATMT_END_DT)) + 1, 1) - 1
    WHERE SUBSTR(T.TACTIC_ID, 8, 3) = '{mne}'
      AND D.APPL_FOR_PROD_TYP IN ('40','41','43')
      AND B.APP_TYP = 'P'
      AND D.PROD_APP_STS_CD IN (32,37,45,47,51,56,62)
      AND D.PROD_APP_DT BETWEEN T.TREATMT_STRT_DT AND T.TREATMT_END_DT
      AND D.PROD_APP_COMPL_DT IS NOT NULL
      AND D.PROD_APP_COMPL_DT >= DATE '2025-01-01'
    ORDER BY D.PROD_APP_COMPL_DT DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM D.PROD_APP_COMPL_DT)  AS yr,
        EXTRACT(MONTH FROM D.PROD_APP_COMPL_DT) AS mo,
        COUNT(DISTINCT A.CLNT_NO)                AS unique_clients,
        COUNT(*)                                 AS total_events,
        D.APPL_FOR_PROD_TYP
    FROM DG6V01.TACTIC_EVNT_IP_AR_HIST   AS T
    JOIN DDWV01.CR_APP_CLNT_RELTN        AS A
        ON T.CLNT_NO = A.CLNT_NO
    JOIN DDWV01.OVRL_CR_APP              AS B
        ON B.CR_APP_ID  = A.CR_APP_ID
        AND B.SYS_SRC_ID = A.SYS_SRC_ID
    JOIN DDWV01.CR_APP_CLNT_PROD_RELTN   AS C
        ON A.CR_APP_ID          = C.CR_APP_ID
        AND A.CR_APP_CLNT_SEQ_NO = C.CR_APP_CLNT_SEQ_NO
        AND A.SYS_SRC_ID        = C.SYS_SRC_ID
    JOIN DDWV01.CR_APP_PROD              AS D
        ON C.CR_APP_ID          = D.CR_APP_ID
        AND C.CR_APP_PROD_SEQ_NO = D.CR_APP_PROD_SEQ_NO
        AND C.SYS_SRC_ID        = D.SYS_SRC_ID
    JOIN DG6V01.ARNGMNT_OWN_HIST        AS E
        ON A.CLNT_NO = E.CLNT_NO
        AND E.MIF_SRVC_ID IN (4)
        AND E.MTH_END_DT = ADD_MONTHS(T.TREATMT_END_DT
            - (EXTRACT(DAY FROM T.TREATMT_END_DT)) + 1, 1) - 1
    WHERE SUBSTR(T.TACTIC_ID, 8, 3) = '{mne}'
      AND D.APPL_FOR_PROD_TYP IN ('40','41','43')
      AND B.APP_TYP = 'P'
      AND D.PROD_APP_STS_CD IN (32,37,45,47,51,56,62)
      AND D.PROD_APP_DT BETWEEN T.TREATMT_STRT_DT AND T.TREATMT_END_DT
      AND D.PROD_APP_COMPL_DT IS NOT NULL
      AND D.PROD_APP_COMPL_DT >= DATE '2025-01-01'
    GROUP BY 1, 2, D.APPL_FOR_PROD_TYP
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- RAT  (RESP Open — DG6V01 arrangement tables) -------------------------

def _rat_mapping(mne):
    return [
        ("clnt_no",            "DG6V01.ARNGMNT_OWN_HIST",    "CLNT_NO"),
        ("acct_no",            "DG6V01.ARNGMNT_HIST",         "AR_ID"),
        ("amount",             "",                              ""),
        ("event_mnemonic_cd",  "",                              mne),
        ("event_type_cd",      "",                              ""),
        ("event_attributes",   "DG6V01.ARNGMNT_HIST",         "JSON(INVSTMT_PLN_TYPE, OP_CLS_STS, MIF_SRVC_ID)"),
        ("event_dt",           "DG6V01.ARNGMNT_HIST",         "DT_OPENED"),
        ("snap_dt",            "",                              "<GENERATED>"),
        ("job_id",             "",                              "<GENERATED>"),
    ]


def rat_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        OWN.CLNT_NO,
        HIST.DT_OPENED         AS SUCCESS_DT,
        HIST.INVSTMT_PLN_TYPE,
        HIST.OP_CLS_STS,
        HIST.AR_ID,
        OWN.MTH_END_DT
    FROM DG6V01.ARNGMNT_OWN_HIST   AS OWN
    INNER JOIN DG6V01.ARNGMNT_HIST  AS HIST
        ON OWN.MIF_ACCT_NO  = HIST.MIF_ACCT_NO
        AND OWN.MIF_SRVC_ID  = HIST.MIF_SRVC_ID
        AND OWN.AR_ID        = HIST.AR_ID
        AND OWN.MTH_END_DT   = HIST.MTH_END_DT
    INNER JOIN DDWV01.AR_BAL_DLY        AS BAL
        ON HIST.AR_ID = BAL.AR_ID
        AND BAL.SNAP_DT = OWN.MTH_END_DT
        AND BAL.SRVC_ID IN (11, 16, 37, 13, 14, 19)
    WHERE HIST.OP_CLS_STS = 'O'
      AND HIST.INVSTMT_PLN_TYPE IN (8)    /* 8 = RESP */
      AND OWN.MTH_END_DT = (SELECT MAX(MTH_END_DT) FROM DG6V01.ARNGMNT_OWN_HIST)
      AND HIST.DT_OPENED >= DATE '2025-01-01'
    ORDER BY HIST.DT_OPENED DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM HIST.DT_OPENED)  AS yr,
        EXTRACT(MONTH FROM HIST.DT_OPENED) AS mo,
        COUNT(DISTINCT OWN.CLNT_NO)        AS unique_clients,
        COUNT(*)                           AS total_events
    FROM DG6V01.ARNGMNT_OWN_HIST   AS OWN
    INNER JOIN DG6V01.ARNGMNT_HIST  AS HIST
        ON OWN.MIF_ACCT_NO  = HIST.MIF_ACCT_NO
        AND OWN.MIF_SRVC_ID  = HIST.MIF_SRVC_ID
        AND OWN.AR_ID        = HIST.AR_ID
        AND OWN.MTH_END_DT   = HIST.MTH_END_DT
    INNER JOIN DDWV01.AR_BAL_DLY        AS BAL
        ON HIST.AR_ID = BAL.AR_ID
        AND BAL.SNAP_DT = OWN.MTH_END_DT
        AND BAL.SRVC_ID IN (11, 16, 37, 13, 14, 19)
    WHERE HIST.OP_CLS_STS = 'O'
      AND HIST.INVSTMT_PLN_TYPE IN (8)    /* 8 = RESP */
      AND OWN.MTH_END_DT = (SELECT MAX(MTH_END_DT) FROM DG6V01.ARNGMNT_OWN_HIST)
      AND HIST.DT_OPENED >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


def rat_campaign(mne):
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        T.CLNT_NO,
        HIST.DT_OPENED         AS SUCCESS_DT,
        HIST.INVSTMT_PLN_TYPE,
        HIST.OP_CLS_STS,
        HIST.AR_ID,
        T.TACTIC_ID,
        T.TREATMT_STRT_DT,
        T.TREATMT_END_DT
    FROM DG6V01.TACTIC_EVNT_IP_AR_HIST AS T
    INNER JOIN DG6V01.ARNGMNT_OWN_HIST  AS OWN
        ON T.CLNT_NO = OWN.CLNT_NO
        AND OWN.MTH_END_DT = ADD_MONTHS(T.TREATMT_END_DT
            - (EXTRACT(DAY FROM T.TREATMT_END_DT)) + 1, 1) - 1
    INNER JOIN DG6V01.ARNGMNT_HIST       AS HIST
        ON OWN.MIF_ACCT_NO  = HIST.MIF_ACCT_NO
        AND OWN.MIF_SRVC_ID  = HIST.MIF_SRVC_ID
        AND OWN.AR_ID        = HIST.AR_ID
        AND OWN.MTH_END_DT   = HIST.MTH_END_DT
    INNER JOIN DDWV01.AR_BAL_DLY        AS BAL
        ON HIST.AR_ID = BAL.AR_ID
        AND BAL.SNAP_DT = OWN.MTH_END_DT
        AND BAL.SRVC_ID IN (11, 16, 37, 13, 14, 19)
    WHERE SUBSTR(T.TACTIC_ID, 8, 3) = '{mne}'
      AND HIST.OP_CLS_STS = 'O'
      AND HIST.INVSTMT_PLN_TYPE IN (8)    /* 8 = RESP */
      AND HIST.DT_OPENED >= T.TREATMT_STRT_DT
      AND HIST.DT_OPENED <= T.TREATMT_END_DT
      AND HIST.DT_OPENED >= DATE '2025-01-01'
    ORDER BY HIST.DT_OPENED DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM HIST.DT_OPENED)  AS yr,
        EXTRACT(MONTH FROM HIST.DT_OPENED) AS mo,
        COUNT(DISTINCT T.CLNT_NO)          AS unique_clients,
        COUNT(*)                           AS total_events
    FROM DG6V01.TACTIC_EVNT_IP_AR_HIST AS T
    INNER JOIN DG6V01.ARNGMNT_OWN_HIST  AS OWN
        ON T.CLNT_NO = OWN.CLNT_NO
        AND OWN.MTH_END_DT = ADD_MONTHS(T.TREATMT_END_DT
            - (EXTRACT(DAY FROM T.TREATMT_END_DT)) + 1, 1) - 1
    INNER JOIN DG6V01.ARNGMNT_HIST       AS HIST
        ON OWN.MIF_ACCT_NO  = HIST.MIF_ACCT_NO
        AND OWN.MIF_SRVC_ID  = HIST.MIF_SRVC_ID
        AND OWN.AR_ID        = HIST.AR_ID
        AND OWN.MTH_END_DT   = HIST.MTH_END_DT
    INNER JOIN DDWV01.AR_BAL_DLY        AS BAL
        ON HIST.AR_ID = BAL.AR_ID
        AND BAL.SNAP_DT = OWN.MTH_END_DT
        AND BAL.SRVC_ID IN (11, 16, 37, 13, 14, 19)
    WHERE SUBSTR(T.TACTIC_ID, 8, 3) = '{mne}'
      AND HIST.OP_CLS_STS = 'O'
      AND HIST.INVSTMT_PLN_TYPE IN (8)    /* 8 = RESP */
      AND HIST.DT_OPENED >= T.TREATMT_STRT_DT
      AND HIST.DT_OPENED <= T.TREATMT_END_DT
      AND HIST.DT_OPENED >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- IDE  (Direct Investing / InvestEase — pre-built success table) ---------

def _ide_mapping(mne):
    return [
        ("clnt_no",            "dl_mr_prod.NBO_IDE_Acquisition",  "CLNT_NO"),
        ("acct_no",            "",                                  ""),
        ("amount",             "",                                  ""),
        ("event_mnemonic_cd",  "",                                  mne),
        ("event_type_cd",      "",                                  ""),
        ("event_attributes",   "dl_mr_prod.NBO_IDE_Acquisition",  "JSON(DI_DT_OPEN, IE_DT_OPEN)"),
        ("event_dt",           "dl_mr_prod.NBO_IDE_Acquisition",  "CASE WHEN DI_DT_OPEN IS NULL THEN IE_DT_OPEN WHEN IE_DT_OPEN IS NULL THEN DI_DT_OPEN WHEN DI_DT_OPEN = IE_DT_OPEN THEN DI_DT_OPEN ELSE NULL END"),
        ("snap_dt",            "",                                  "<GENERATED>"),
        ("job_id",             "",                                  "<GENERATED>"),
    ]

def ide_organic():
    dt = ("CASE WHEN DI_DT_OPEN IS NULL THEN IE_DT_OPEN "
          "WHEN IE_DT_OPEN IS NULL THEN DI_DT_OPEN "
          "WHEN DI_DT_OPEN = IE_DT_OPEN THEN DI_DT_OPEN "
          "ELSE NULL END")
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        {dt} AS SUCCESS_DT,
        DI_DT_OPEN,
        IE_DT_OPEN,
        Control,
        Success
    FROM dl_mr_prod.NBO_IDE_Acquisition
    WHERE Success = 1
      AND {dt} >= DATE '2025-01-01'
    ORDER BY 2 DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM {dt})  AS yr,
        EXTRACT(MONTH FROM {dt}) AS mo,
        COUNT(DISTINCT CLNT_NO)   AS unique_clients,
        COUNT(*)                  AS total_events
    FROM dl_mr_prod.NBO_IDE_Acquisition
    WHERE Success = 1
      AND {dt} >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary

def ide_campaign(mne):
    dt = ("CASE WHEN DI_DT_OPEN IS NULL THEN IE_DT_OPEN "
          "WHEN IE_DT_OPEN IS NULL THEN DI_DT_OPEN "
          "WHEN DI_DT_OPEN = IE_DT_OPEN THEN DI_DT_OPEN "
          "ELSE NULL END")
    sample = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        {dt} AS SUCCESS_DT,
        DI_DT_OPEN,
        IE_DT_OPEN,
        Control,
        Treatmt_strt_dt,
        Success
    FROM dl_mr_prod.NBO_IDE_Acquisition
    WHERE Control = 'Action'
      AND Success = 1
      AND {dt} >= DATE '2025-01-01'
    ORDER BY 2 DESC

);
disconnect from teradata;
quit;"""

    summary = f"""\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM {dt})  AS yr,
        EXTRACT(MONTH FROM {dt}) AS mo,
        COUNT(DISTINCT CLNT_NO)   AS unique_clients,
        COUNT(*)                  AS total_events
    FROM dl_mr_prod.NBO_IDE_Acquisition
    WHERE Control = 'Action'
      AND Success = 1
      AND {dt} >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- GIS  (GIC Open — pre-built success table) -----------------------------

def _gis_mapping(mne):
    return [
        ("clnt_no",            "dl_mr_prod.NBO_GIC_Acquisition",  "CLNT_NO"),
        ("acct_no",            "",                                  ""),
        ("amount",             "",                                  ""),
        ("event_mnemonic_cd",  "",                                  mne),
        ("event_type_cd",      "",                                  ""),
        ("event_attributes",   "dl_mr_prod.NBO_GIC_Acquisition",  ""),
        ("event_dt",           "dl_mr_prod.NBO_GIC_Acquisition",  "dt_open"),
        ("snap_dt",            "",                                  "<GENERATED>"),
        ("job_id",             "",                                  "<GENERATED>"),
    ]

def gis_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        dt_open               AS SUCCESS_DT,
        Control,
        Success
    FROM dl_mr_prod.NBO_GIC_Acquisition
    WHERE Success = 1
      AND dt_open >= DATE '2025-01-01'
    ORDER BY dt_open DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM dt_open)  AS yr,
        EXTRACT(MONTH FROM dt_open) AS mo,
        COUNT(DISTINCT CLNT_NO)     AS unique_clients,
        COUNT(*)                    AS total_events
    FROM dl_mr_prod.NBO_GIC_Acquisition
    WHERE Success = 1
      AND dt_open >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary

def gis_campaign(mne):
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        dt_open               AS SUCCESS_DT,
        Control,
        Treatmt_strt_dt,
        Success
    FROM dl_mr_prod.NBO_GIC_Acquisition
    WHERE Control = 'Action'
      AND Success = 1
      AND dt_open >= DATE '2025-01-01'
    ORDER BY dt_open DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM dt_open)  AS yr,
        EXTRACT(MONTH FROM dt_open) AS mo,
        COUNT(DISTINCT CLNT_NO)     AS unique_clients,
        COUNT(*)                    AS total_events
    FROM dl_mr_prod.NBO_GIC_Acquisition
    WHERE Control = 'Action'
      AND Success = 1
      AND dt_open >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# --- TAO  (Registered Account Open — pre-built success table) --------------

def _tao_mapping(mne):
    return [
        ("clnt_no",            "dl_mr_prod.NBO_TAO_Acquisition",  "CLNT_NO"),
        ("acct_no",            "dl_mr_prod.NBO_TAO_Acquisition",  "PLN_AR_ID"),
        ("amount",             "",                                  ""),
        ("event_mnemonic_cd",  "",                                  mne),
        ("event_type_cd",      "",                                  ""),
        ("event_attributes",   "dl_mr_prod.NBO_TAO_Acquisition",  ""),
        ("event_dt",           "dl_mr_prod.NBO_TAO_Acquisition",  "dt_open"),
        ("snap_dt",            "",                                  "<GENERATED>"),
        ("job_id",             "",                                  "<GENERATED>"),
    ]

def tao_organic():
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        dt_open               AS SUCCESS_DT,
        PLN_AR_ID,
        Control
    FROM dl_mr_prod.NBO_TAO_Acquisition
    WHERE PLN_AR_ID IS NOT NULL
      AND dt_open >= DATE '2025-01-01'
    ORDER BY dt_open DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM dt_open)  AS yr,
        EXTRACT(MONTH FROM dt_open) AS mo,
        COUNT(DISTINCT CLNT_NO)     AS unique_clients,
        COUNT(*)                    AS total_events
    FROM dl_mr_prod.NBO_TAO_Acquisition
    WHERE PLN_AR_ID IS NOT NULL
      AND dt_open >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary

def tao_campaign(mne):
    sample = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT TOP 10
        CLNT_NO,
        dt_open               AS SUCCESS_DT,
        PLN_AR_ID,
        Control,
        Treatmt_strt_dt
    FROM dl_mr_prod.NBO_TAO_Acquisition
    WHERE Control = 'Action'
      AND PLN_AR_ID IS NOT NULL
      AND dt_open >= DATE '2025-01-01'
    ORDER BY dt_open DESC

);
disconnect from teradata;
quit;"""

    summary = """\
proc sql;
%connectsql
select * from connection to teradata (

    SELECT
        EXTRACT(YEAR FROM dt_open)  AS yr,
        EXTRACT(MONTH FROM dt_open) AS mo,
        COUNT(DISTINCT CLNT_NO)     AS unique_clients,
        COUNT(*)                    AS total_events
    FROM dl_mr_prod.NBO_TAO_Acquisition
    WHERE Control = 'Action'
      AND PLN_AR_ID IS NOT NULL
      AND dt_open >= DATE '2025-01-01'
    GROUP BY 1, 2
    ORDER BY 1, 2

);
disconnect from teradata;
quit;"""
    return sample, summary


# ---------------------------------------------------------------------------
# Tab configuration
# (mnemonic, title, mapping_fn, organic_fn, campaign_fn)
# ---------------------------------------------------------------------------
TABS = [
    ("VCN", "Card Acquisition",         _card_acq_mapping,   card_acq_organic,   card_acq_campaign),
    ("VDA", "Card Acquisition",         _card_acq_mapping,   card_acq_organic,   card_acq_campaign),
    ("VDT", "Card Activation",          _card_actv_mapping,  card_actv_organic,  card_actv_campaign),
    ("VUI", "Card Usage",               _card_usage_mapping, card_usage_organic, card_usage_campaign),
    ("VUT", "Wallet Provisioning",      _wallet_mapping,     wallet_organic,     wallet_campaign),
    ("VAW", "Wallet Provisioning",      _wallet_mapping,     wallet_organic,     wallet_campaign),
    ("IRI", "Intl Money Transfer",      _imt_mapping,        imt_organic,        imt_campaign),
    ("O2P", "Overdraft Open",           _o2p_mapping,        o2p_organic,        o2p_campaign),
    ("RAT", "RESP Open",               _rat_mapping,        rat_organic,        rat_campaign),
    ("IDE", "Direct Investing Open",   _ide_mapping,        ide_organic,        ide_campaign),
    ("GIS", "GIC Open",               _gis_mapping,        gis_organic,        gis_campaign),
    ("TAO", "Registered Acct Open",   _tao_mapping,        tao_organic,        tao_campaign),
]


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
def _write_sql_block(ws, row, sql_text):
    """Write SQL lines into column A starting at `row`. Returns next empty row."""
    for line in sql_text.splitlines():
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = SQL_FONT
        cell.alignment = WRAP_ALIGN
        row += 1
    return row


def build_workbook(output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for mne, title, mapping_fn, organic_fn, campaign_fn in TABS:
        ws = wb.create_sheet(title=f"{mne}_Success")
        ws.column_dimensions["A"].width = COL_A_WIDTH
        ws.column_dimensions["B"].width = COL_B_WIDTH
        ws.column_dimensions["C"].width = COL_C_WIDTH

        row = 1

        # ==== SOURCE-TO-TARGET MAPPING TABLE ================================
        for col, hdr in enumerate(["Target Column", "source table", "source column/logic"], 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.font = MAP_HDR_FONT
            c.fill = MAP_HDR_FILL
            c.border = THIN_BORDER
        row += 1

        for tcol, stbl, slogic in mapping_fn(mne):
            ws.cell(row=row, column=1, value=tcol).font = MAP_DATA_FONT
            ws.cell(row=row, column=2, value=stbl).font = MAP_DATA_FONT
            ws.cell(row=row, column=3, value=slogic).font = MAP_DATA_FONT
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        row += 2  # blank

        # ==== ORGANIC SECTION ===============================================
        ws.cell(row=row, column=1, value="Query:").font = LABEL_FONT
        row += 1

        # -- Organic Sample --
        c = ws.cell(row=row, column=1, value="ORGANIC — Sample (10 rows)")
        c.font = HEADER_FONT
        c.fill = GREEN_FILL
        row += 1

        org_sample, org_summary = organic_fn()
        row = _write_sql_block(ws, row, org_sample)
        row += 1

        # -- Organic Summary --
        c = ws.cell(row=row, column=1, value="ORGANIC — Summary (by year-month)")
        c.font = HEADER_FONT
        c.fill = YELLOW_FILL
        row += 1

        row = _write_sql_block(ws, row, org_summary)
        row += 2

        # ==== CAMPAIGN SECTION ==============================================

        # -- Campaign Sample --
        c = ws.cell(row=row, column=1, value=f"CAMPAIGN ({mne}) — Sample (10 rows)")
        c.font = HEADER_FONT
        c.fill = BLUE_FILL
        row += 1

        cmp_sample, cmp_summary = campaign_fn(mne)
        row = _write_sql_block(ws, row, cmp_sample)
        row += 1

        # -- Campaign Summary --
        c = ws.cell(row=row, column=1, value=f"CAMPAIGN ({mne}) — Summary (by year-month)")
        c.font = HEADER_FONT
        c.fill = YELLOW_FILL
        row += 1

        row = _write_sql_block(ws, row, cmp_summary)

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    print(f"Tabs: {[s.title for s in wb.worksheets]}")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    OUTPUT = (
        "/mnt/c/Users/andre/New_projects/"
        "NBA Souccess Library - Copy/metadata/success_queries.xlsx"
    )
    build_workbook(OUTPUT)
