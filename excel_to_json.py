#!/usr/bin/env python3
"""
excel_to_json.py - Success Library Intake Processor (Schema v2.0)

Processes Excel intake forms and updates the metadata JSON.
Supports the enhanced Schema v2.0 with technical, measurement, and governance blocks.

Features:
- Processes all .xlsx files in intake/pending/
- Creates backup before any changes
- Validates entries and checks for conflicts
- Shows summary and asks for confirmation
- Moves processed files to intake/processed/
- Generates code file stubs from template

Usage:
    python3 excel_to_json.py

Schema v2.0 Additions:
- technical: source, table_path, date_field, client_field, partition_field, filters
- measurement: default_window_days, attribution_model
- governance: status, last_validated
- complexity: SIMPLE, TRANSFORM, MULTI_TABLE

Version: 2.0
Updated: 2026-01-24
"""

import json
import os
import shutil
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install --user openpyxl")
    exit(1)

# Configuration
METADATA_PATH = Path("metadata/success_library_index.json")
PENDING_DIR = Path("intake/pending")
PROCESSED_DIR = Path("intake/processed")
BACKUP_DIR = Path("metadata/backups")
CODE_DIR = Path("code")
CODE_TEMPLATE_PATH = Path("intake/template/code_template.md")

# ============================================================
# COLUMN DEFINITIONS - Schema v2.0
# ============================================================

# Sheet 1: Business Metadata (original columns + version)
BUSINESS_COLUMNS = [
    "action",           # new, update
    "metric_id",        # VVD_ACQ_001
    "metric_name",      # Card Acquisition
    "product",          # VVD, CC, MTG
    "line_of_business", # Debit Cards, Lending
    "metric_type",      # Acquisition, Activation, Usage, Retention
    "pillar",           # Conversion, Engagement, Retention, Profitability
    "business_definition",  # Plain English
    "source_tables",    # Comma-separated list
    "staged_source",    # Staged table if applicable
    "grain",            # Client, Account, Transaction
    "owner",            # Marketing Analytics
    "campaigns_using",  # VCN, VDA, VDT
]

# Sheet 2: Technical Metadata (new in v2.0)
TECHNICAL_COLUMNS = [
    "metric_id",            # Must match Sheet 1
    "complexity",           # SIMPLE, TRANSFORM, MULTI_TABLE
    "source",               # HIVE, EDW, TERADATA
    "table_path",           # Full HDFS path (null for EDW)
    "date_field",           # ISS_DT, TXN_DT, ACTV_DT
    "client_field",         # CLNT_NO (output field name)
    "partition_field",      # CAPTR_DT, SNAP_DT
    "client_extraction_logic",  # SUBSTR(CLNT_CRD_NO, 7, 9) if needed
    "add_card_type",        # TRUE/FALSE
    # Filters as separate columns (key=value pairs)
    "filter_1_field",       # e.g., STS_CD
    "filter_1_condition",   # e.g., IN
    "filter_1_value",       # e.g., 06,08
    "filter_2_field",
    "filter_2_condition",
    "filter_2_value",
    "filter_3_field",
    "filter_3_condition",
    "filter_3_value",
    "filter_4_field",
    "filter_4_condition",
    "filter_4_value",
    # Join info (for MULTI_TABLE complexity)
    "join_table",           # Secondary table for join
    "join_type",            # INNER, LEFT
    "join_condition",       # A.KEY = B.KEY
]

# Sheet 3: Measurement & Governance (new in v2.0)
MEASUREMENT_COLUMNS = [
    "metric_id",            # Must match Sheet 1
    "default_window_days",  # 90, 180, 365
    "attribution_model",    # FIRST_TOUCH, LAST_TOUCH, LINEAR
    "status",               # ACTIVE, PLACEHOLDER, DEPRECATED, RETIRED
    "last_validated",       # YYYY-MM-DD or blank
    "notes",                # Any additional notes
]

# Valid values for validation
VALID_COMPLEXITY = ["SIMPLE", "TRANSFORM", "MULTI_TABLE"]
VALID_SOURCE = ["HIVE", "EDW", "TERADATA"]
VALID_ATTRIBUTION = ["FIRST_TOUCH", "LAST_TOUCH", "LINEAR", "TIME_DECAY"]
VALID_STATUS = ["ACTIVE", "PLACEHOLDER", "DEPRECATED", "RETIRED"]
VALID_ACTIONS = ["new", "update"]
VALID_PILLARS = ["Conversion", "Engagement", "Retention", "Profitability", "Share of Wallet"]
VALID_GRAINS = ["Client", "Account", "Transaction"]


def read_json(path):
    """Read and parse JSON file."""
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def write_json(path, data):
    """Write data to JSON file with pretty formatting."""
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def create_backup(source_path, backup_dir):
    """Create a timestamped backup of the JSON file."""
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"success_library_index_backup_{timestamp}.json"
    backup_path = backup_dir / backup_name
    shutil.copy2(source_path, backup_path)
    return backup_path


def parse_list_field(value):
    """Parse comma-separated string into list."""
    if not value:
        return []
    if isinstance(value, list):
        return value
    return [item.strip() for item in str(value).split(",") if item.strip()]


def parse_int_field(value):
    """Parse integer field, return None if empty."""
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def parse_bool_field(value):
    """Parse boolean field."""
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    return str(value).upper() in ["TRUE", "YES", "1", "Y"]


def clean_value(value):
    """Clean up a cell value."""
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    return value


def read_sheet_as_dict(ws, columns, start_row=3):
    """
    Read a worksheet into a list of dicts.
    Returns dict keyed by metric_id for easy lookup.
    """
    result = {}
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=start_row, max_col=len(columns)), start=start_row):
        values = [cell.value for cell in row]

        # Skip empty rows
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        # Skip example rows
        if len(values) > 0 and values[0] and "example" in str(values[0]).lower():
            continue

        entry = {}
        for i, col_name in enumerate(columns):
            value = values[i] if i < len(values) else None
            entry[col_name] = clean_value(value)

        # Get metric_id for keying
        metric_id = entry.get("metric_id")
        if not metric_id:
            # For business sheet, metric_id might not be first
            if "action" in entry and entry.get("action"):
                errors.append(f"Row {row_num}: 'metric_id' is required")
            continue

        entry["_source_row"] = row_num
        result[metric_id] = entry

    return result, errors


def build_filters_dict(technical_entry):
    """Build a filters object from the filter columns."""
    filters = {}

    for i in range(1, 5):
        field = technical_entry.get(f"filter_{i}_field")
        condition = technical_entry.get(f"filter_{i}_condition")
        value = technical_entry.get(f"filter_{i}_value")

        if not field:
            continue

        # Parse the condition and value into appropriate structure
        if condition and condition.upper() == "IN":
            # Multiple values, comma-separated
            filters[field] = parse_list_field(value)
        elif condition and condition.upper() == "NOT_NULL":
            filters[f"{field}_NOT_NULL"] = True
        elif condition and condition.upper() in ["=", "==", "EQ"]:
            # Try to parse as int, else keep as string
            parsed = parse_int_field(value)
            filters[field] = parsed if parsed is not None else value
        elif condition and condition.upper() in [">", "GT"]:
            filters[f"{field}_GT"] = parse_int_field(value) or value
        elif condition and condition.upper() in ["<", "LT"]:
            filters[f"{field}_LT"] = parse_int_field(value) or value
        else:
            # Default: just store the value
            parsed = parse_int_field(value)
            filters[field] = parsed if parsed is not None else value

    return filters if filters else None


def build_technical_block(technical_entry):
    """Build the technical block from technical sheet entry."""
    if not technical_entry:
        return None

    block = {
        "source": technical_entry.get("source") or "HIVE",
        "table_path": technical_entry.get("table_path"),
        "date_field": technical_entry.get("date_field"),
        "client_field": technical_entry.get("client_field") or "CLNT_NO",
    }

    # Optional fields
    if technical_entry.get("partition_field"):
        block["partition_field"] = technical_entry["partition_field"]

    if technical_entry.get("client_extraction_logic"):
        block["client_extraction_logic"] = technical_entry["client_extraction_logic"]

    if parse_bool_field(technical_entry.get("add_card_type")):
        block["add_card_type"] = True

    # Build filters
    filters = build_filters_dict(technical_entry)
    if filters:
        block["filters"] = filters

    # Join info (for MULTI_TABLE)
    if technical_entry.get("join_table"):
        block["join"] = {
            "table": technical_entry["join_table"],
            "type": technical_entry.get("join_type") or "INNER",
            "on": technical_entry.get("join_condition"),
        }

    return block


def build_measurement_block(measurement_entry):
    """Build the measurement block from measurement sheet entry."""
    if not measurement_entry:
        return {
            "default_window_days": 90,
            "attribution_model": "FIRST_TOUCH"
        }

    return {
        "default_window_days": parse_int_field(measurement_entry.get("default_window_days")) or 90,
        "attribution_model": measurement_entry.get("attribution_model") or "FIRST_TOUCH"
    }


def build_governance_block(measurement_entry):
    """Build the governance block from measurement sheet entry."""
    if not measurement_entry:
        return {
            "status": "PLACEHOLDER",
            "last_validated": None
        }

    return {
        "status": measurement_entry.get("status") or "PLACEHOLDER",
        "last_validated": measurement_entry.get("last_validated")
    }


def read_intake_file(filepath):
    """
    Read an intake Excel file with multiple sheets.
    Returns combined entries and any errors.
    """
    entries = []
    errors = []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # ====================
    # Sheet 1: Business Metadata
    # ====================
    business_sheet = None
    for name in ["business metadata", "data entry", "business", "metadata"]:
        if name in sheet_names_lower:
            business_sheet = wb[sheet_names_lower[name]]
            break

    # Fallback to first non-instructions sheet
    if business_sheet is None:
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() not in ["instructions", "technical", "measurement"]:
                business_sheet = wb[sheet_name]
                break

    if business_sheet is None:
        business_sheet = wb.active

    business_data, biz_errors = read_sheet_as_dict(business_sheet, BUSINESS_COLUMNS)
    errors.extend([f"Business Sheet: {e}" for e in biz_errors])

    # ====================
    # Sheet 2: Technical Metadata (optional)
    # ====================
    technical_data = {}
    for name in ["technical metadata", "technical", "tech"]:
        if name in sheet_names_lower:
            tech_sheet = wb[sheet_names_lower[name]]
            technical_data, tech_errors = read_sheet_as_dict(tech_sheet, TECHNICAL_COLUMNS)
            errors.extend([f"Technical Sheet: {e}" for e in tech_errors])
            break

    # ====================
    # Sheet 3: Measurement & Governance (optional)
    # ====================
    measurement_data = {}
    for name in ["measurement", "measurement & governance", "governance"]:
        if name in sheet_names_lower:
            meas_sheet = wb[sheet_names_lower[name]]
            measurement_data, meas_errors = read_sheet_as_dict(meas_sheet, MEASUREMENT_COLUMNS)
            errors.extend([f"Measurement Sheet: {e}" for e in meas_errors])
            break

    wb.close()

    # ====================
    # Combine all sheets by metric_id
    # ====================
    for metric_id, biz_entry in business_data.items():
        action = biz_entry.get("action")

        if not action:
            continue

        if action not in VALID_ACTIONS:
            errors.append(f"Row {biz_entry['_source_row']}: 'action' must be 'new' or 'update', got '{action}'")
            continue

        # Parse list fields
        biz_entry["source_tables"] = parse_list_field(biz_entry.get("source_tables"))
        biz_entry["campaigns_using"] = parse_list_field(biz_entry.get("campaigns_using"))

        # Get technical data for this metric
        tech_entry = technical_data.get(metric_id)

        # Get measurement data for this metric
        meas_entry = measurement_data.get(metric_id)

        # Build combined entry
        combined = {
            "action": action,
            "metric_id": metric_id,
            "metric_name": biz_entry.get("metric_name"),
            "product": biz_entry.get("product"),
            "line_of_business": biz_entry.get("line_of_business"),
            "metric_type": biz_entry.get("metric_type"),
            "pillar": biz_entry.get("pillar"),
            "business_definition": biz_entry.get("business_definition"),
            "source_tables": biz_entry["source_tables"],
            "staged_source": biz_entry.get("staged_source") or "",
            "grain": biz_entry.get("grain"),
            "owner": biz_entry.get("owner"),
            "campaigns_using": biz_entry["campaigns_using"],
            # v2.0 blocks
            "complexity": tech_entry.get("complexity") if tech_entry else "SIMPLE",
            "technical": build_technical_block(tech_entry),
            "measurement": build_measurement_block(meas_entry),
            "governance": build_governance_block(meas_entry),
            # Tracking
            "_source_file": filepath.name,
            "_source_row": biz_entry["_source_row"],
        }

        entries.append(combined)

    return entries, errors


def validate_entries(entries, existing_metrics):
    """Validate entries against existing data and schema rules."""
    errors = []
    warnings = []

    existing_ids = {m["metric_id"] for m in existing_metrics}
    seen_ids = set()

    for entry in entries:
        metric_id = entry["metric_id"]
        action = entry["action"]
        source = f"{entry['_source_file']} row {entry['_source_row']}"

        # Check for duplicates within intake
        if metric_id in seen_ids:
            errors.append(f"{source}: Duplicate metric_id '{metric_id}' in intake files")
            continue
        seen_ids.add(metric_id)

        # Validate action vs existence
        if action == "new" and metric_id in existing_ids:
            errors.append(f"{source}: metric_id '{metric_id}' already exists. Use 'update' instead.")
        elif action == "update" and metric_id not in existing_ids:
            errors.append(f"{source}: metric_id '{metric_id}' not found. Use 'new' instead.")

        # Validate required fields for new entries
        if action == "new":
            required_for_new = ["metric_name", "product", "metric_type", "pillar"]
            for field in required_for_new:
                if not entry.get(field):
                    errors.append(f"{source}: '{field}' is required for new metrics")

        # Validate technical block if present
        tech = entry.get("technical")
        if tech:
            if tech.get("source") and tech["source"] not in VALID_SOURCE:
                warnings.append(f"{source}: source '{tech['source']}' not in {VALID_SOURCE}")

            if not tech.get("date_field"):
                warnings.append(f"{source}: 'date_field' is recommended in technical block")

        # Validate complexity
        complexity = entry.get("complexity")
        if complexity and complexity not in VALID_COMPLEXITY:
            warnings.append(f"{source}: complexity '{complexity}' not in {VALID_COMPLEXITY}")

        # Validate governance
        gov = entry.get("governance", {})
        if gov.get("status") and gov["status"] not in VALID_STATUS:
            warnings.append(f"{source}: status '{gov['status']}' not in {VALID_STATUS}")

    return errors, warnings


def apply_changes(entries, library_data):
    """Apply intake entries to library data. Returns summary of changes."""
    metrics = library_data.get("metrics", [])
    metrics_by_id = {m["metric_id"]: m for m in metrics}

    added = []
    updated = []

    for entry in entries:
        action = entry["action"]
        metric_id = entry["metric_id"]

        # Remove internal tracking fields
        clean_entry = {k: v for k, v in entry.items() if not k.startswith("_")}
        del clean_entry["action"]
        del clean_entry["complexity"]  # This goes in technical block conceptually

        if action == "new":
            # Generate code_path
            product = clean_entry.get("product", "UNKNOWN")
            clean_entry["code_path"] = f"code/{product}/{metric_id}.md"
            clean_entry["version"] = "v1.0"

            # Ensure all fields exist
            for field in ["line_of_business", "business_definition", "staged_source", "owner"]:
                if field not in clean_entry or clean_entry[field] is None:
                    clean_entry[field] = ""

            # Ensure v2.0 blocks exist
            if not clean_entry.get("technical"):
                clean_entry["technical"] = {
                    "source": "HIVE",
                    "table_path": None,
                    "date_field": None,
                    "client_field": "CLNT_NO",
                    "filters": {}
                }

            if not clean_entry.get("measurement"):
                clean_entry["measurement"] = {
                    "default_window_days": 90,
                    "attribution_model": "FIRST_TOUCH"
                }

            if not clean_entry.get("governance"):
                clean_entry["governance"] = {
                    "status": "PLACEHOLDER",
                    "last_validated": None
                }

            metrics.append(clean_entry)
            added.append(metric_id)

        elif action == "update":
            existing = metrics_by_id[metric_id]
            changes = []

            for field, value in clean_entry.items():
                # Skip empty values (keep existing)
                if value is None or value == "":
                    continue
                # Skip empty lists (keep existing)
                if isinstance(value, list) and len(value) == 0:
                    continue
                # Handle nested dicts (technical, measurement, governance)
                if isinstance(value, dict) and field in existing and isinstance(existing[field], dict):
                    # Merge nested dicts
                    for k, v in value.items():
                        if v is not None and v != "" and v != existing[field].get(k):
                            old_val = existing[field].get(k)
                            existing[field][k] = v
                            changes.append(f"{field}.{k}: '{old_val}' → '{v}'")
                elif value != existing.get(field):
                    old_val = existing.get(field)
                    existing[field] = value
                    changes.append(f"{field}: '{old_val}' → '{value}'")

            if changes:
                updated.append((metric_id, changes))

    library_data["metrics"] = metrics
    library_data["library_info"]["last_updated"] = datetime.now().strftime("%Y-%m-%d")

    # Update schema version to 2.0 if not already
    if library_data.get("schema_version") != "2.0":
        library_data["schema_version"] = "2.0"
        library_data["schema_notes"] = "Added 'technical', 'measurement', and 'governance' blocks per metric."

    return added, updated


def generate_code_stub(metric_entry, template_path, output_dir):
    """Generate a code file stub from the template."""
    if not template_path.exists():
        return None

    # Read template
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()

    # Basic replacements
    replacements = {
        "{METRIC_ID}": metric_entry.get("metric_id", ""),
        "{Metric Name}": metric_entry.get("metric_name", ""),
        "{Owner Name}": metric_entry.get("owner", ""),
        "{Product code: VVD, CC, MTG, etc.}": metric_entry.get("product", ""),
        "{Acquisition / Activation / Usage / Provisioning / Retention}": metric_entry.get("metric_type", ""),
        "{Conversion / Engagement / Retention / Profitability / Share of Wallet}": metric_entry.get("pillar", ""),
        "{Campaign codes, comma-separated}": ", ".join(metric_entry.get("campaigns_using", [])),
        "{Client / Account / Transaction}": metric_entry.get("grain", ""),
    }

    content = template
    for old, new in replacements.items():
        content = content.replace(old, str(new) if new else "")

    # Write to output
    product = metric_entry.get("product", "UNKNOWN")
    metric_id = metric_entry["metric_id"]
    output_path = output_dir / product / f"{metric_id}.md"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Only write if file doesn't exist
    if not output_path.exists():
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return output_path

    return None


def move_to_processed(filepath):
    """Move processed file to processed directory with timestamp."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = filepath.stem
    new_name = f"{stem}_processed_{timestamp}.xlsx"
    dest = PROCESSED_DIR / new_name
    shutil.move(str(filepath), str(dest))
    return dest


def main():
    print("=" * 60)
    print("Success Library - Intake Processor (Schema v2.0)")
    print("=" * 60)
    print()

    # Check for pending files
    PENDING_DIR.mkdir(parents=True, exist_ok=True)
    pending_files = list(PENDING_DIR.glob("*.xlsx"))

    if not pending_files:
        print(f"No intake files found in {PENDING_DIR}/")
        print("Place Excel files in this folder and run again.")
        print()
        print("Expected Excel structure:")
        print("  Sheet 1: 'Business Metadata' - metric_id, metric_name, product, ...")
        print("  Sheet 2: 'Technical' (optional) - source, table_path, filters, ...")
        print("  Sheet 3: 'Measurement' (optional) - window_days, status, ...")
        return

    print(f"Found {len(pending_files)} intake file(s):")
    for f in pending_files:
        print(f"  - {f.name}")
    print()

    # Read all intake files
    all_entries = []
    all_errors = []

    for filepath in pending_files:
        print(f"Reading {filepath.name}...")
        entries, errors = read_intake_file(filepath)
        all_entries.extend(entries)
        all_errors.extend(errors)
        print(f"  Found {len(entries)} entries")

    print()

    if all_errors:
        print("ERRORS found in intake files:")
        for error in all_errors:
            print(f"  ✗ {error}")
        print()
        print("Please fix these errors and run again.")
        return

    if not all_entries:
        print("No valid entries found in intake files.")
        return

    # Load existing data
    print(f"Loading {METADATA_PATH}...")
    library_data = read_json(METADATA_PATH)
    existing_metrics = library_data.get("metrics", [])
    print(f"  Current metrics: {len(existing_metrics)}")
    print()

    # Validate
    print("Validating entries...")
    errors, warnings = validate_entries(all_entries, existing_metrics)

    if errors:
        print("VALIDATION ERRORS:")
        for error in errors:
            print(f"  ✗ {error}")
        print()
        print("Please fix these errors and run again.")
        return

    if warnings:
        print("Warnings (non-blocking):")
        for warning in warnings:
            print(f"  ! {warning}")
        print()

    print("Validation passed.")
    print()

    # Preview changes
    added, updated = apply_changes(
        [e.copy() for e in all_entries],
        json.loads(json.dumps(library_data))
    )

    print("=" * 60)
    print("SUMMARY OF CHANGES")
    print("=" * 60)
    print()

    if added:
        print(f"NEW METRICS ({len(added)}):")
        for metric_id in added:
            print(f"  + {metric_id}")
        print()

    if updated:
        print(f"UPDATED METRICS ({len(updated)}):")
        for metric_id, changes in updated:
            print(f"  ~ {metric_id}")
            for change in changes:
                print(f"      {change}")
        print()

    if not added and not updated:
        print("No changes to apply.")
        return

    print("=" * 60)
    print()

    # Confirm
    response = input("Apply these changes? (yes/no): ").strip().lower()

    if response != "yes":
        print("Cancelled. No changes made.")
        return

    print()

    # Create backup
    print("Creating backup...")
    backup_path = create_backup(METADATA_PATH, BACKUP_DIR)
    print(f"  Backup saved: {backup_path}")

    # Apply changes for real
    print("Applying changes...")
    library_data = read_json(METADATA_PATH)  # Re-read fresh
    apply_changes(all_entries, library_data)
    write_json(METADATA_PATH, library_data)
    print(f"  Updated: {METADATA_PATH}")

    # Generate code stubs for new metrics
    if added:
        print("Generating code file stubs...")
        for entry in all_entries:
            if entry["metric_id"] in added:
                stub_path = generate_code_stub(entry, CODE_TEMPLATE_PATH, CODE_DIR)
                if stub_path:
                    print(f"  Created: {stub_path}")
                else:
                    print(f"  Skipped (exists): code/{entry['product']}/{entry['metric_id']}.md")

    # Move processed files
    print("Moving processed files...")
    for filepath in pending_files:
        dest = move_to_processed(filepath)
        print(f"  {filepath.name} → {dest.name}")

    print()
    print("=" * 60)
    print("DONE!")
    print("=" * 60)
    print()
    print("Next steps:")
    print("  1. Edit code files in code/{product}/{metric_id}.md")
    print("     - Add SQL query in 'SQL (Data Warehouse)' section")
    print("     - Add PySpark code in 'PySpark (Hive)' section")
    print("  2. Run: python3 build.py")
    print()


if __name__ == "__main__":
    main()
